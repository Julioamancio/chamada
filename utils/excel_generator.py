"""
Gerador de planilhas Excel profissionais para relatórios
Utiliza openpyxl para criar arquivos Excel com formatação avançada
"""
import os
import io
from datetime import datetime
from typing import Dict, Any, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows


class ExcelGenerator:
    """Gerador de planilhas Excel profissionais com múltiplas abas e formatação"""
    
    # Cores do tema
    COLORS = {
        'primary': '1476F2',
        'success': '28A745',
        'danger': 'DC3545',
        'warning': 'FFC107',
        'info': '17A2B8',
        'light': 'F8F9FA',
        'dark': '343A40'
    }
    
    @staticmethod
    def gerar_excel(dados: Dict[str, Any]) -> bytes:
        """Gera arquivo Excel completo com múltiplas abas"""
        wb = Workbook()
        
        # Remover aba padrão
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Criar abas
        ExcelGenerator._criar_aba_resumo(wb, dados)
        ExcelGenerator._criar_aba_ranking(wb, dados)
        ExcelGenerator._criar_aba_detalhado(wb, dados)
        ExcelGenerator._criar_aba_estatisticas(wb, dados)
        
        # Salvar em buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    @staticmethod
    def _criar_aba_resumo(wb: Workbook, dados: Dict[str, Any]):
        """Cria aba com resumo executivo"""
        ws = wb.create_sheet("📊 Resumo Executivo")
        
        # Cabeçalho principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"RELATÓRIO DE CHAMADA ESCOLAR - {dados['turma']['nome']}"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=ExcelGenerator.COLORS['primary'], 
                                   end_color=ExcelGenerator.COLORS['primary'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Informações gerais
        ws['A3'] = "Gerado em:"
        ws['B3'] = datetime.now().strftime('%d/%m/%Y às %H:%M')
        ws['A4'] = "Período:"
        periodo = dados.get('periodo', {})
        if periodo.get('inicio') and periodo.get('fim'):
            ws['B4'] = f"{periodo['inicio'].strftime('%d/%m/%Y')} a {periodo['fim'].strftime('%d/%m/%Y')}"
        else:
            ws['B4'] = "Todos os registros"
        
        # Estatísticas principais
        stats = dados['estatisticas_gerais']
        
        estatisticas = [
            ("📚 Total de Alunos", stats['total_alunos']),
            ("📋 Total de Chamadas", stats['total_chamadas']),
            ("✅ Total de Presenças", stats['total_presencas']),
            ("❌ Total de Faltas", stats['total_faltas']),
            ("📊 Presença Geral", f"{stats['percentual_presenca_geral']:.1f}%"),
            ("🎯 Média da Turma", f"{stats['media_pontuacao']:.1f} pts"),
            ("🥇 Maior Pontuação", f"{stats['maior_pontuacao']:.1f} pts"),
            ("🥉 Menor Pontuação", f"{stats['menor_pontuacao']:.1f} pts")
        ]
        
        # Adicionar estatísticas com formatação
        row = 6
        for descricao, valor in estatisticas:
            ws[f'A{row}'] = descricao
            ws[f'B{row}'] = valor
            
            # Formatação da descrição
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'A{row}'].fill = PatternFill(start_color=ExcelGenerator.COLORS['light'], 
                                           end_color=ExcelGenerator.COLORS['light'], fill_type='solid')
            
            # Formatação do valor
            if isinstance(valor, str) and '%' in str(valor):
                ws[f'B{row}'].font = Font(color=ExcelGenerator.COLORS['success'], bold=True)
            else:
                ws[f'B{row}'].font = Font(color=ExcelGenerator.COLORS['primary'], bold=True)
            
            row += 1
        
        # Top 5 alunos
        ws[f'A{row+1}'] = "🏆 TOP 5 ALUNOS"
        ws[f'A{row+1}'].font = Font(size=14, bold=True, color=ExcelGenerator.COLORS['primary'])
        
        # Cabeçalhos do top 5
        headers = ['Pos.', 'Aluno', 'Pontuação', '% Presença']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row+3, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=ExcelGenerator.COLORS['primary'],
                                   end_color=ExcelGenerator.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados do top 5
        top_5 = dados.get('ranking', [])[:5]
        for i, aluno in enumerate(top_5):
            row_num = row + 4 + i
            ws[f'A{row_num}'] = aluno['posicao']
            ws[f'B{row_num}'] = aluno['aluno']
            ws[f'C{row_num}'] = round(aluno['pontuacao_total'], 1)
            ws[f'D{row_num}'] = f"{aluno['percentual_presenca']:.1f}%"
            
            # Formatação especial para os 3 primeiros
            if i < 3:
                colors = [ExcelGenerator.COLORS['warning'], 
                         ExcelGenerator.COLORS['light'], 
                         ExcelGenerator.COLORS['warning']]
                for col in range(1, 5):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = PatternFill(start_color=colors[i], end_color=colors[i], fill_type='solid')
                    if i == 1:  # Prata
                        cell.font = Font(bold=True)
        
        # Ajustar larguras das colunas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    @staticmethod
    def _criar_aba_ranking(wb: Workbook, dados: Dict[str, Any]):
        """Cria aba com ranking completo dos alunos"""
        ws = wb.create_sheet("🏆 Ranking Completo")
        
        # Cabeçalho
        headers = ['Posição', 'Aluno', 'Pontuação Total', 'Presenças', 'Faltas', '% Presença']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=ExcelGenerator.COLORS['primary'],
                                   end_color=ExcelGenerator.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Dados do ranking
        for row, aluno in enumerate(dados.get('ranking', []), 2):
            ws[f'A{row}'] = aluno['posicao']
            ws[f'B{row}'] = aluno['aluno']
            ws[f'C{row}'] = round(aluno['pontuacao_total'], 1)
            ws[f'D{row}'] = aluno['presencas']
            ws[f'E{row}'] = aluno['faltas']
            ws[f'F{row}'] = round(aluno['percentual_presenca'], 1)
            
            # Formatação condicional para as 3 primeiras posições
            if aluno['posicao'] <= 3:
                colors = {1: ExcelGenerator.COLORS['warning'], 
                         2: 'C0C0C0', 
                         3: 'CD7F32'}
                fill_color = colors.get(aluno['posicao'], ExcelGenerator.COLORS['light'])
                
                for col in range(1, 7):
                    cell = ws.cell(row=row, column=col)
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
                    if aluno['posicao'] == 1:
                        cell.font = Font(bold=True)
        
        # Formatação condicional para percentual de presença
        if len(dados.get('ranking', [])) > 1:
            # Regra de cores para percentual de presença
            rule = ColorScaleRule(start_type='min', start_color='FFEB9C',
                                 mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                 end_type='max', end_color='63BE7B')
            ws.conditional_formatting.add(f'F2:F{len(dados["ranking"])+1}', rule)
            
            # Barra de dados para pontuação
            rule = DataBarRule(start_type='min', end_type='max', color=ExcelGenerator.COLORS['primary'])
            ws.conditional_formatting.add(f'C2:C{len(dados["ranking"])+1}', rule)
        
        # Ajustar larguras das colunas
        column_widths = [10, 25, 15, 12, 12, 12]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64+i)].width = width
        
        # Criar gráfico de barras
        if len(dados.get('ranking', [])) > 0:
            chart = BarChart()
            chart.title = "Pontuação dos Alunos"
            chart.y_axis.title = "Pontuação"
            chart.x_axis.title = "Alunos"
            
            # Limitar a 10 primeiros para melhor visualização
            max_rows = min(11, len(dados['ranking']) + 1)
            data = Reference(ws, min_col=3, min_row=1, max_row=max_rows, max_col=3)
            cats = Reference(ws, min_col=2, min_row=2, max_row=max_rows)
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            
            ws.add_chart(chart, "H2")
    
    @staticmethod
    def _criar_aba_detalhado(wb: Workbook, dados: Dict[str, Any]):
        """Cria aba com detalhamento por etapa"""
        ws = wb.create_sheet("📋 Detalhado por Etapa")
        
        # Cabeçalhos
        headers = ['Aluno']
        etapas = dados.get('etapas', [])
        
        for etapa in etapas:
            headers.append(etapa['nome'])
        headers.append('TOTAL')
        
        # Adicionar cabeçalhos
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=ExcelGenerator.COLORS['primary'],
                                   end_color=ExcelGenerator.COLORS['primary'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Adicionar dados
        for row, aluno in enumerate(dados.get('resultados_alunos', []), 2):
            ws[f'A{row}'] = aluno['aluno']
            
            col = 2
            for etapa in etapas:
                pontuacao = aluno['pontuacao_por_etapa'].get(etapa['id'], 0)
                ws.cell(row=row, column=col, value=round(pontuacao, 1))
                col += 1
            
            # Total
            ws.cell(row=row, column=col, value=round(aluno['pontuacao_total'], 1))
            ws.cell(row=row, column=col).font = Font(bold=True)
        
        # Formatação condicional para valores
        if len(dados.get('resultados_alunos', [])) > 1:
            # Aplicar formatação condicional nas colunas de pontuação
            for col in range(2, len(headers) + 1):
                col_letter = chr(64 + col)
                range_address = f'{col_letter}2:{col_letter}{len(dados["resultados_alunos"])+1}'
                
                rule = ColorScaleRule(start_type='min', start_color='FFEB9C',
                                     mid_type='percentile', mid_value=50, mid_color='FFFF00',
                                     end_type='max', end_color='63BE7B')
                ws.conditional_formatting.add(range_address, rule)
        
        # Ajustar larguras das colunas
        ws.column_dimensions['A'].width = 25
        for col in range(2, len(headers) + 1):
            ws.column_dimensions[chr(64 + col)].width = 12
        
        # Adicionar totais na última linha
        if dados.get('resultados_alunos'):
            total_row = len(dados['resultados_alunos']) + 3
            ws[f'A{total_row}'] = 'MÉDIA DA TURMA'
            ws[f'A{total_row}'].font = Font(bold=True)
            ws[f'A{total_row}'].fill = PatternFill(start_color=ExcelGenerator.COLORS['light'],
                                                  end_color=ExcelGenerator.COLORS['light'], fill_type='solid')
            
            # Calcular médias
            col = 2
            for etapa in etapas:
                total_etapa = sum(aluno['pontuacao_por_etapa'].get(etapa['id'], 0) 
                                for aluno in dados['resultados_alunos'])
                media = total_etapa / len(dados['resultados_alunos']) if dados['resultados_alunos'] else 0
                
                cell = ws.cell(row=total_row, column=col, value=round(media, 1))
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=ExcelGenerator.COLORS['light'],
                                       end_color=ExcelGenerator.COLORS['light'], fill_type='solid')
                col += 1
            
            # Média total
            cell = ws.cell(row=total_row, column=col, value=round(dados['estatisticas_gerais']['media_pontuacao'], 1))
            cell.font = Font(bold=True, color=ExcelGenerator.COLORS['primary'])
            cell.fill = PatternFill(start_color=ExcelGenerator.COLORS['light'],
                                   end_color=ExcelGenerator.COLORS['light'], fill_type='solid')
    
    @staticmethod
    def _criar_aba_estatisticas(wb: Workbook, dados: Dict[str, Any]):
        """Cria aba com estatísticas avançadas"""
        ws = wb.create_sheet("📊 Estatísticas Avançadas")
        
        # Título
        ws.merge_cells('A1:D1')
        ws['A1'] = "ANÁLISE ESTATÍSTICA AVANÇADA"
        ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=ExcelGenerator.COLORS['primary'],
                                   end_color=ExcelGenerator.COLORS['primary'], fill_type='solid')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        stats = dados['estatisticas_gerais']
        
        # Seção 1: Estatísticas Gerais
        ws['A3'] = "📊 ESTATÍSTICAS GERAIS"
        ws['A3'].font = Font(size=12, bold=True, color=ExcelGenerator.COLORS['primary'])
        
        estatisticas_gerais = [
            ("Total de Alunos", stats['total_alunos']),
            ("Total de Chamadas Realizadas", stats['total_chamadas']),
            ("Total de Presenças Registradas", stats['total_presencas']),
            ("Total de Faltas Registradas", stats['total_faltas']),
            ("Percentual Geral de Presença", f"{stats['percentual_presenca_geral']:.2f}%"),
            ("Taxa de Comparecimento", f"{stats['percentual_presenca_geral']:.2f}%")
        ]
        
        row = 4
        for desc, valor in estatisticas_gerais:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = valor
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Seção 2: Estatísticas de Pontuação
        row += 1
        ws[f'A{row}'] = "🎯 ESTATÍSTICAS DE PONTUAÇÃO"
        ws[f'A{row}'].font = Font(size=12, bold=True, color=ExcelGenerator.COLORS['primary'])
        
        estatisticas_pontuacao = [
            ("Média da Turma", f"{stats['media_pontuacao']:.2f} pontos"),
            ("Maior Pontuação Individual", f"{stats['maior_pontuacao']:.2f} pontos"),
            ("Menor Pontuação Individual", f"{stats['menor_pontuacao']:.2f} pontos"),
            ("Amplitude (Maior - Menor)", f"{stats['maior_pontuacao'] - stats['menor_pontuacao']:.2f} pontos")
        ]
        
        row += 1
        for desc, valor in estatisticas_pontuacao:
            ws[f'A{row}'] = desc
            ws[f'B{row}'] = valor
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # Seção 3: Análise de Distribuição
        if dados.get('resultados_alunos'):
            pontuacoes = [aluno['pontuacao_total'] for aluno in dados['resultados_alunos']]
            percentuais = [aluno['percentual_presenca'] for aluno in dados['resultados_alunos']]
            
            # Calcular quartis
            pontuacoes_sorted = sorted(pontuacoes)
            n = len(pontuacoes_sorted)
            
            q1_idx = n // 4
            q2_idx = n // 2
            q3_idx = 3 * n // 4
            
            q1 = pontuacoes_sorted[q1_idx] if q1_idx < n else 0
            q2 = pontuacoes_sorted[q2_idx] if q2_idx < n else 0
            q3 = pontuacoes_sorted[q3_idx] if q3_idx < n else 0
            
            row += 1
            ws[f'A{row}'] = "📈 ANÁLISE DE DISTRIBUIÇÃO"
            ws[f'A{row}'].font = Font(size=12, bold=True, color=ExcelGenerator.COLORS['primary'])
            
            distribuicao = [
                ("1º Quartil (25%)", f"{q1:.2f} pontos"),
                ("2º Quartil (50% - Mediana)", f"{q2:.2f} pontos"),
                ("3º Quartil (75%)", f"{q3:.2f} pontos"),
                ("Desvio Padrão (aproximado)", f"{(max(pontuacoes) - min(pontuacoes)) / 4:.2f}"),
                ("Alunos Acima da Média", sum(1 for p in pontuacoes if p > stats['media_pontuacao'])),
                ("Alunos Abaixo da Média", sum(1 for p in pontuacoes if p < stats['media_pontuacao']))
            ]
            
            row += 1
            for desc, valor in distribuicao:
                ws[f'A{row}'] = desc
                ws[f'B{row}'] = valor
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
        
        # Ajustar larguras das colunas
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15