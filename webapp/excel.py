from __future__ import annotations

from typing import BinaryIO, List
from werkzeug.datastructures import FileStorage


def _strip_header(names: List[str]) -> List[str]:
    if names and names[0].strip().lower() in {"nome", "name", "aluno"}:
        return names[1:]
    return names


def names_from_xlsx_stream(stream: BinaryIO) -> List[str]:
    from openpyxl import load_workbook

    wb = load_workbook(filename=stream, read_only=True, data_only=True)
    ws = wb.active
    names: List[str] = []
    # iter_rows works in read_only mode; take only first column
    for (value,) in ws.iter_rows(min_col=1, max_col=1, values_only=True):
        val = str(value).strip() if value is not None else ""
        if val:
            names.append(val)
    return _strip_header(names)


def names_from_csv_stream(stream: BinaryIO) -> List[str]:
    import csv
    # Ensure text mode with utf-8
    import io

    if isinstance(stream, io.BytesIO) or hasattr(stream, "read"):
        text = io.TextIOWrapper(stream, encoding="utf-8", newline="")
    else:
        text = stream  # assume text file-like
    reader = csv.reader(text)
    names: List[str] = []
    for row in reader:
        if not row:
            continue
        val = str(row[0]).strip()
        if val:
            names.append(val)
    return _strip_header(names)


def names_from_upload(file: FileStorage) -> List[str]:
    filename = (file.filename or "").lower()
    if filename.endswith(".xlsx"):
        # Reset stream in case it's been read
        file.stream.seek(0)
        return names_from_xlsx_stream(file.stream)
    if filename.endswith(".csv"):
        file.stream.seek(0)
        return names_from_csv_stream(file.stream)
    # Fallback: treat as plain text, one name per line
    file.stream.seek(0)
    content = file.stream.read()
    try:
        text = content.decode("utf-8")
    except AttributeError:
        text = str(content)
    names = [line.strip() for line in text.splitlines() if line.strip()]
    return _strip_header(names)


# Backwards compatibility: older imports may reference this name
def names_from_first_column(file) -> List[str]:
    """Compatibility alias. Accepts the same upload object and delegates.

    If a raw stream is passed, it will be treated as XLSX by default.
    """
    try:
        # If this is a Werkzeug FileStorage, delegate to the main handler
        from werkzeug.datastructures import FileStorage as _FS

        if isinstance(file, _FS):
            return names_from_upload(file)
    except Exception:
        pass
    # Fallback: assume it's an XLSX stream
    try:
        return names_from_xlsx_stream(file)
    except Exception:
        # Last resort: try CSV
        return names_from_csv_stream(file)
